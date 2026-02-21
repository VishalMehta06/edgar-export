import pandas as pd
from io import StringIO
from datetime import datetime, timedelta
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

from app.Client import Client
from app.logger import get_logger

logger = get_logger(__name__)


class Stock:
	"""
	Organize the results for a single stock unit.
	"""

	def __init__(self, client: Client, ticker: str, 
			  	 filing_forms: list[str] = ["10-K", "10-Q"],
				 years: int = 10):
		"""		
		:param ticker: The ticker of the stock. Capitalization doesn't matter.
		:type ticker: str
		:param client: The client to be used to fetch data.
		:type client: Client
		:param filing_forms: A list of strings of filing 'form's to track.
		:type filing_forms: list[str]
		:param years: Only include filings from the last N years. Defaults to 10.
		:type years: int
		"""
		self.ticker = ticker.upper()
		self.client = client
		self.years = years
		logger.info("Initializing Stock for ticker=%s (last %d years)", 
					self.ticker, self.years)
		self.cik = client.get_cik(ticker)

		if not self.cik:
			logger.error(
				"Could not resolve CIK for ticker=%s — filings will be empty",
				self.ticker
			)

		self.filings = self._init_filings(filing_forms)
		logger.info(
			"Stock initialized for ticker=%s: CIK=%s, filings=%d",
			self.ticker, self.cik, len(self.filings)
		)

	def _init_filings(self, filing_forms: list[str]) -> list[dict]:
		"""
		Initialize self.filings. Only includes filings within the last
		`self.years` years. The cutoff is pushed down into Client.get_filings
		so that pagination stops early rather than fetching historical data
		that will never be displayed.

		:param filing_forms: A list of strings of filing 'form's to track.
		:type filing_forms: list[str]
		"""
		cutoff_date = datetime.now() - timedelta(days=365 * self.years)
		logger.debug(
			"Fetching filings with cutoff=%s and forms=%s for ticker=%s",
			cutoff_date.date(), filing_forms, self.ticker
		)

		# Pass cutoff_date into get_filings so it stops paginating as soon
		# as it reaches filings older than the window — no wasted requests.
		all_filings = self.client.get_filings(self.cik, cutoff_date=cutoff_date)
		logger.debug(
			"Received %d filings within cutoff for ticker=%s",
			len(all_filings), self.ticker
		)

		selected_filings = []

		for filing in all_filings:
			# Date filtering already done by get_filings — only filter by form
			if filing["form"] not in filing_forms:
				continue

			accn = filing["accn"]
			try:
				data = self.client.get_filing_data(self.cik, accn)
				selected_filings.append({
					"metadata": filing,
					"reports": data
				})
				logger.debug(
					"Added filing accn=%s form=%s for ticker=%s",
					accn, filing["form"], self.ticker
				)
			except Exception as e:
				status_code = getattr(getattr(e, "response", None), "status_code", None)
				if status_code == 404:
					logger.info(
						"Skipping filing accn=%s form=%s for ticker=%s: 404 Not Found",
						accn, filing["form"], self.ticker
					)
				else:
					logger.warning(
						"Skipping filing accn=%s form=%s for ticker=%s: %s",
						accn, filing["form"], self.ticker, e
					)

		logger.info(
			"_init_filings complete for ticker=%s: %d/%d filings selected",
			self.ticker, len(selected_filings), len(all_filings)
		)
		return selected_filings
	
	def export_url(self, url: str, filename: str, 
				   report_category: str = "statement") -> None:
		"""
		Export a report from a URL to excel.

		:param url: The URL of the report
		:type url: str
		:param filename: The output filename
		:type filename: str
		:param report_category: The category of the report. Either "document", 
		"statement", or "disclosure". 
		:type report_category: str 
		"""
		logger.info(
			"Exporting URL=%s to file=%s (category=%s)", 
			url, filename, report_category
		)
		resp = self.client._fetch_response(url)

		if not resp:
			logger.error("Failed to fetch report URL=%s for export", url)
			raise RuntimeError(f"Could not fetch report from {url}")

		soup = BeautifulSoup(resp.text, "lxml")
		tables = self._extract_tables(soup)
		text = self._extract_text_blocks(soup)

		# Remove Garbage XBRL Tables
		before = len(tables)
		tables = [t for t in tables if not self._is_xbrl_table(t[1])]
		logger.debug(
			"Removed %d XBRL table(s), %d table(s) remaining",
			before - len(tables), len(tables)
		)

		wb = Workbook()
		wb.remove(wb.active)

		for sheet_name, df in tables:
			ws = wb.create_sheet(sheet_name[:31])
			for row in dataframe_to_rows(df, index=False, header=True):
				ws.append(row)

		if report_category != "statement":
			ws_text = wb.create_sheet("Text_Content")
			for row in dataframe_to_rows(text, index=False, header=True):
				ws_text.append(row)

		wb.save(filename)
		logger.info("Export complete: saved to %s", filename)

	def _is_xbrl_table(self, df):
		EXPECTED_FIRST_COLUMN = [
			"Name:",
			"Namespace Prefix:",
			"Data Type:",
			"Balance Type:",
			"Period Type:"
		]
		return df.iloc[:, 0].tolist() == EXPECTED_FIRST_COLUMN

	def _extract_tables(self, soup):
		"""
		Helper method to extract all tables from html reports.
		"""
		tables = soup.find_all("table")
		extracted_tables = []

		for idx, table in enumerate(tables, start=1):
			try:
				df = pd.read_html(StringIO(str(table)))[0]
				extracted_tables.append((f"Table_{idx}", df))
			except ValueError:
				logger.debug("Skipping unparseable table at index %d", idx)
				continue

		logger.debug("Extracted %d table(s) from report", len(extracted_tables))
		return extracted_tables
	
	def _extract_text_blocks(self, soup):
		"""
		Helper method to extract all text blocks from html reports.
		"""
		text_rows = []

		for tag in soup.find_all(["h1", "h2", "h3", "h4", "p", "div"]):
			text = tag.get_text(strip=True)
			if text:
				text_rows.append({
					"Tag": tag.name,
					"Text": text
				})

		logger.debug("Extracted %d text block(s) from report", len(text_rows))
		return pd.DataFrame(text_rows)
