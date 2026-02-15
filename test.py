import requests
import pandas as pd
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

URL = "https://www.sec.gov/Archives/edgar/data/1090872/000109087225000087/R19.htm"
HEADERS = {
    "User-Agent": "Vishal Mehta, vvmehta06@gmail.com"
}

OUTPUT_FILE = "SEC_Filing_Export.xlsx"


def fetch_html(url):
    response = requests.get(url, headers=HEADERS)
    response.raise_for_status()
    return response.text


def extract_tables(soup):
    tables = soup.find_all("table")
    extracted_tables = []

    for idx, table in enumerate(tables, start=1):
        try:
            df = pd.read_html(str(table))[0]
            extracted_tables.append((f"Table_{idx}", df))
        except ValueError:
            # Skip tables that pandas can't parse
            continue

    return extracted_tables


def extract_text_blocks(soup):
    text_rows = []

    for tag in soup.find_all(["h1", "h2", "h3", "h4", "p", "div"]):
        text = tag.get_text(strip=True)
        if text:
            text_rows.append({
                "Tag": tag.name,
                "Text": text
            })

    return pd.DataFrame(text_rows)


def write_to_excel(tables, text_df, output_file):
    wb = Workbook()
    wb.remove(wb.active)

    # Write text content
    ws_text = wb.create_sheet("Text_Content")
    for row in dataframe_to_rows(text_df, index=False, header=True):
        ws_text.append(row)

    # Write each table to its own sheet
    for sheet_name, df in tables:
        ws = wb.create_sheet(sheet_name[:31])
        for row in dataframe_to_rows(df, index=False, header=True):
            ws.append(row)

    wb.save(output_file)


def main():
    html = fetch_html(URL)
    soup = BeautifulSoup(html, "lxml")

    tables = extract_tables(soup)
    text_df = extract_text_blocks(soup)

    write_to_excel(tables, text_df, OUTPUT_FILE)
    print(f"Export completed: {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
